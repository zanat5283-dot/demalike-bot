import asyncio
import os
import re
import sqlite3
from datetime import datetime
from io import BytesIO
import aiohttp
from aiogram import Bot, Dispatcher, types, F
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import BufferedInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

try:
    from PIL import Image

    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# ================= OCR & API НАСТРОЙКИ =================
OCR_SPACE_API_KEY = os.environ.get("OCR_SPACE_API_KEY", "K84914005788957")
OCR_SPACE_URL = "https://api.ocr.space/parse/image"
OCR_AVAILABLE = bool(OCR_SPACE_API_KEY)

# Ссылка на развернутое Google Apps Script Web App
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbylyoio9cqsISsZR9VrAyMQ5vYtV1BystGObMwKO2nspxAjQx8eoIviac5Y5qgFiRHw/exec"

BOT_TOKEN = os.environ.get("BOT_TOKEN", "8846227339:AAHSRBu5bC_HWCMNSKUSHFryMXkhfcdPLh8")
ADMIN_IDS = [7810711826]  # Telegram ID админа
TARGET_CHAT_IDS = [-1003924194210]
KASPI_PAY_LINK = "https://pay.kaspi.kz/pay/lbjevwyf"

bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode="Markdown"))
dp = Dispatcher(storage=MemoryStorage())
scheduler = AsyncIOScheduler(timezone="Asia/Almaty")

DB_PATH = "demalike_payments.db"


# ================= ИНТЕГРАЦИЯ С GOOGLE TABLES =================
async def sync_to_google_sheet(full_name: str, phone: str = "", notes: str = "", klass: str = "Продленка",
                               group: str = "DemaLike"):
    """Отправляет запись в Google Таблицу через веб-приложение Apps Script."""
    if not GOOGLE_SCRIPT_URL:
        return

    payload = {
        "name": full_name or "Не указано",
        "klass": klass,
        "group": group,
        "parentName": full_name or "Не указано",
        "parentPhone": phone or "Не указан",
        "notes": notes
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(GOOGLE_SCRIPT_URL, json=payload, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                if resp.status == 200:
                    print(f"[Google Sheets] Успешно синхронизировано: {full_name}")
                else:
                    print(f"[Google Sheets Error] Статус ответа: {resp.status}")
    except Exception as e:
        print(f"[Google Sheets Exception]: {e}")


# ================= БАЗА ДАННЫХ =================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            phone_number TEXT,
            full_name TEXT,
            chat_id INTEGER
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            month TEXT NOT NULL,
            amount REAL NOT NULL,
            pay_type TEXT NOT NULL,
            paid_at TEXT NOT NULL
        )
    """)

    cursor.execute("PRAGMA table_info(users)")
    existing_columns = {row[1] for row in cursor.fetchall()}
    required_columns = {
        "username": "TEXT",
        "phone_number": "TEXT",
        "full_name": "TEXT",
        "chat_id": "INTEGER",
    }
    for col_name, col_type in required_columns.items():
        if col_name not in existing_columns:
            cursor.execute(f"ALTER TABLE users ADD COLUMN {col_name} {col_type}")

    conn.commit()
    conn.close()


def upsert_user_profile(user_id: int, username: str = "", phone_number: str = "",
                        full_name: str = "", chat_id: int = 0):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    clean_username = username.lstrip("@").lower() if username else ""

    cursor.execute("""
        INSERT INTO users (user_id, username, phone_number, full_name, chat_id)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            username=COALESCE(NULLIF(?, ''), username),
            phone_number=COALESCE(NULLIF(?, ''), phone_number),
            full_name=COALESCE(NULLIF(?, ''), full_name),
            chat_id=COALESCE(NULLIF(?, 0), chat_id)
    """, (user_id, clean_username, phone_number, full_name, chat_id,
          clean_username, phone_number, full_name, chat_id))
    conn.commit()
    conn.close()


def get_user_profile(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT full_name, phone_number, username FROM users WHERE user_id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"full_name": row[0], "phone_number": row[1], "username": row[2]}
    return None


def record_payment(user_id: int, amount: float, pay_type: str, month: str = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    month = month or datetime.now().strftime("%Y-%m")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute(
        "INSERT INTO payments (user_id, month, amount, pay_type, paid_at) VALUES (?, ?, ?, ?, ?)",
        (user_id, month, amount, pay_type, now)
    )
    conn.commit()
    conn.close()


def find_user_id_by_username(username: str):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    clean_username = username.lstrip("@").lower()
    cursor.execute("SELECT user_id FROM users WHERE LOWER(username) = ?", (clean_username,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else None


def get_month_report(month: str = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    month = month or datetime.now().strftime("%Y-%m")
    cursor.execute("""
        SELECT u.user_id, u.full_name, u.username, p.amount, p.pay_type, p.paid_at
        FROM payments p
        JOIN users u ON u.user_id = p.user_id
        WHERE p.month = ?
        ORDER BY u.full_name, u.user_id, p.paid_at
    """, (month,))
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_unpaid_list(month: str = None):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    month = month or datetime.now().strftime("%Y-%m")
    cursor.execute("""
        SELECT u.user_id, u.full_name, u.username
        FROM users u
        WHERE u.user_id NOT IN (
            SELECT user_id FROM payments WHERE month = ?
        )
    """, (month,))
    rows = cursor.fetchall()
    conn.close()
    return rows


def build_month_excel_report(month: str) -> BytesIO:
    rows = get_month_report(month)

    wb = Workbook()
    ws = wb.active
    ws.title = month[:31]

    headers = ["Имя", "Юзернейм", "Сумма (₸)", "Способ оплаты", "Дата и время оплаты"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    payers = set()
    current_user_id = None
    person_total = 0.0
    person_count = 0
    person_label = ("", "")

    def flush_person_subtotal():
        if person_count > 1:
            idx = ws.max_row + 1
            ws.append(
                ["", "", person_total, f"Итого по {person_label[0] or person_label[1]}", f"{person_count} чек(а/ов)"])
            for cell in ws[idx]:
                cell.font = Font(italic=True)

    for user_id, full_name, username, amount, pay_type, paid_at in rows:
        if current_user_id is not None and user_id != current_user_id:
            flush_person_subtotal()
            person_total, person_count = 0.0, 0

        current_user_id = user_id
        person_label = (full_name, f"@{username}" if username else "")
        person_total += amount
        person_count += 1
        payers.add(user_id)

        ws.append([
            full_name or "—",
            f"@{username}" if username else "—",
            amount,
            pay_type,
            paid_at,
        ])
        total += amount

    flush_person_subtotal()

    ws.append([])
    total_row_idx = ws.max_row + 1
    ws.append(["", "", total, "ИТОГО ЗА МЕСЯЦ", f"{len(payers)} чел. / {len(rows)} чек(ов)"])
    for cell in ws[total_row_idx]:
        cell.font = Font(bold=True)

    column_widths = [24, 20, 14, 20, 22]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ================= РАСПОЗНАВАНИЕ СУММЫ НА ЧЕКЕ (OCR) =================
GROUPED_THOUSANDS = re.compile(r'\b\d{1,3}(?:[ \u00a0]\d{3})+\b')
AMOUNT_WITH_CURRENCY = re.compile(r'(\d[\d\s\u00a0]{0,9})\s*(?:₸|тг\.?|kzt|тенге|F\b)', re.IGNORECASE)
AMOUNT_WITH_KEYWORD = re.compile(r'(?:сумма|итого|к оплате|перевод)\D{0,15}?(\d[\d\s\u00a0]{1,9})', re.IGNORECASE)


def extract_amount_from_text(text: str):
    grouped_matches = GROUPED_THOUSANDS.findall(text)
    if grouped_matches:
        raw = grouped_matches[0].replace(" ", "").replace("\u00a0", "")
        try:
            value = float(raw)
            if 100 <= value <= 10_000_000:
                return value
        except ValueError:
            pass

    candidates = []
    for regex in (AMOUNT_WITH_CURRENCY, AMOUNT_WITH_KEYWORD):
        for m in regex.finditer(text):
            raw = m.group(1).replace(" ", "").replace("\u00a0", "")
            try:
                value = float(raw)
            except ValueError:
                continue
            if 10 <= value <= 10_000_000:
                candidates.append(value)

    if not candidates:
        return None
    return max(candidates)


MAX_OCR_FILE_BYTES = 1_000_000


def compress_image_for_ocr(image_bytes: bytes) -> bytes:
    if len(image_bytes) <= MAX_OCR_FILE_BYTES or not PIL_AVAILABLE:
        return image_bytes

    image = Image.open(BytesIO(image_bytes)).convert("RGB")
    quality = 85
    while True:
        buf = BytesIO()
        image.save(buf, format="JPEG", quality=quality)
        data = buf.getvalue()
        if len(data) <= MAX_OCR_FILE_BYTES or (quality <= 40 and max(image.size) <= 800):
            return data
        if quality > 40:
            quality -= 15
        else:
            w, h = image.size
            image = image.resize((int(w * 0.7), int(h * 0.7)))


async def ocr_via_api(file_bytes: bytes, is_pdf: bool) -> str:
    form = aiohttp.FormData()
    form.add_field("apikey", OCR_SPACE_API_KEY)
    form.add_field("language", "rus")
    form.add_field("scale", "true")
    form.add_field("OCREngine", "2")
    if is_pdf:
        form.add_field("file", file_bytes, filename="receipt.pdf", content_type="application/pdf")
    else:
        form.add_field("file", file_bytes, filename="receipt.jpg", content_type="image/jpeg")

    async with aiohttp.ClientSession() as session:
        async with session.post(OCR_SPACE_URL, data=form, timeout=aiohttp.ClientTimeout(total=30)) as resp:
            result = await resp.json(content_type=None)

    if result.get("IsErroredOnProcessing"):
        raise RuntimeError(result.get("ErrorMessage") or "OCR.space processing error")

    parsed_results = result.get("ParsedResults") or []
    return "\n".join(r.get("ParsedText", "") for r in parsed_results)


async def extract_amount_from_telegram_file(file_id: str, is_pdf: bool):
    if not OCR_AVAILABLE:
        return None
    try:
        tg_file = await bot.get_file(file_id)
        file_io = await bot.download_file(tg_file.file_path)
        file_bytes = file_io.read()

        if not is_pdf:
            file_bytes = compress_image_for_ocr(file_bytes)
        elif len(file_bytes) > MAX_OCR_FILE_BYTES:
            print("[OCR] PDF больше 1 МБ — пропускаем автораспознавание")
            return None

        text = await ocr_via_api(file_bytes, is_pdf=is_pdf)
        return extract_amount_from_text(text)
    except Exception as e:
        print(f"[OCR Error]: {e}")
        return None


def format_amount(amount: float) -> str:
    return f"{amount:,.0f}".replace(",", " ")


# ================= FSM СОСТОЯНИЯ =================
class AdminCashState(StatesGroup):
    waiting_for_contact = State()
    waiting_for_amount = State()


class AdminUsernameState(StatesGroup):
    waiting_for_username = State()
    waiting_for_amount = State()


class AdminApproveState(StatesGroup):
    waiting_for_amount = State()


# ================= КЛАВИАТУРЫ =================
def get_group_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.button(text="💳 Оплатить через Kaspi")
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)


def get_admin_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.button(text="💵 Провести оплату (Наличные)")
    builder.button(text="👤 Отметить по юзернейму")
    builder.button(text="📊 Отчёт за месяц")
    builder.button(text="📥 Excel-отчёт")
    builder.button(text="📋 Список должников")
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)


def get_contact_request_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.button(text="📱 Поделиться контактом", request_contact=True)
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True, one_time_keyboard=True)


def parse_amount(text: str):
    cleaned = text.strip().replace(" ", "").replace(",", ".")
    try:
        value = float(cleaned)
        if value <= 0:
            return None
        return value
    except ValueError:
        return None


def escape_md(text: str) -> str:
    if not text:
        return text
    for ch in ("_", "*", "`", "["):
        text = text.replace(ch, "\\" + ch)
    return text


# ================= ЛИЧНЫЕ СООБЩЕНИЯ (АДМИНКА) =================

@dp.message(F.chat.type == "private", Command("testcheck"))
async def admin_test_notification(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    report_lines = [f"🧪 **Тест уведомлений** — {datetime.now().strftime('%H:%M:%S')}\n"]
    report_lines.append(f"ADMIN_IDS в коде: `{ADMIN_IDS}`")
    report_lines.append(f"Ваш реальный ID: `{message.from_user.id}`")
    report_lines.append("✅ Совпадает" if message.from_user.id in ADMIN_IDS else "❌ НЕ совпадает — вот и причина!")
    report_lines.append(f"\nЦелевые группы TARGET_CHAT_IDS: `{TARGET_CHAT_IDS}`")
    report_lines.append(f"OCR-автораспознавание суммы: {'✅ включено' if OCR_AVAILABLE else '❌ выключено'}\n")

    for admin_id in ADMIN_IDS:
        builder = InlineKeyboardBuilder()
        builder.button(text="✅ Подтвердить", callback_data="approve_test")
        builder.button(text="❌ Отклонить", callback_data="reject_test")
        builder.adjust(2)

        test_caption = (
            f"📥 **Новый чек на проверку (DemaLike)!**\n\n"
            f"• **Имя:** {escape_md(message.from_user.full_name)}\n"
            f"• **Профиль:** тест\n"
            f"• **ID:** `0`\n"
            f"• **Группа:** `тест`"
        )
        try:
            await bot.send_message(admin_id, test_caption, reply_markup=builder.as_markup())
            report_lines.append(f"➡️ Отправка на ID `{admin_id}`: ✅ УСПЕШНО")
        except Exception as e:
            report_lines.append(f"➡️ Отправка на ID `{admin_id}`: ❌ ОШИБКА\n`{type(e).__name__}: {e}`")

    await message.answer("\n".join(report_lines), parse_mode=None)


@dp.message(F.chat.type == "private", F.text == "💵 Провести оплату (Наличные)")
async def admin_cash_start(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.set_state(AdminCashState.waiting_for_contact)
    await message.answer(
        "📲 **Отправьте контакт человека:**\n"
        "Перешлите карточку контакта нужного клиента в этот чат.",
        reply_markup=get_contact_request_keyboard()
    )


@dp.message(F.chat.type == "private", F.text == "👤 Отметить по юзернейму")
async def admin_username_start(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        return
    await state.set_state(AdminUsernameState.waiting_for_username)
    await message.answer(
        "✏️ **Введите Telegram-юзернейм пользователя:**\n"
        "Например: `@username` или просто `username`",
        reply_markup=types.ReplyKeyboardRemove()
    )


@dp.message(F.chat.type == "private", F.text == "📥 Excel-отчёт")
async def admin_excel_report(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    month = datetime.now().strftime("%Y-%m")
    rows = get_month_report(month)

    if not rows:
        await message.answer(f"📥 За **{month}** оплат ещё нет — Excel-файл собирать не из чего.")
        return

    total = sum(r[3] for r in rows)
    unique_payers = len({r[0] for r in rows})
    buffer = build_month_excel_report(month)
    filename = f"demalike_report_{month}.xlsx"

    await message.answer_document(
        BufferedInputFile(buffer.read(), filename=filename),
        caption=(
            f"📥 **Excel-отчёт за {month}**\n\n"
            f"💰 **Итого собрано:** {format_amount(total)} ₸\n"
            f"👥 **Оплатили:** {unique_payers} чел. ({len(rows)} чек(ов))"
        )
    )


@dp.message(F.chat.type == "private")
async def private_chat_router(message: types.Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer(
            "ℹ️ Личный чат бота предназначен только для администрации.\n"
            "Все операции выполняются **в рабочей группе DemaLike**."
        )
        return

    if message.text in ["/start", "/admin"]:
        await state.clear()
        await message.answer(
            "🔐 **Панель Администратора DemaLike**\n\n"
            "ℹ️ Если уведомления о чеках не приходят — отправьте `/testcheck`.",
            reply_markup=get_admin_keyboard()
        )
        return

    current_state = await state.get_state()

    if current_state == AdminCashState.waiting_for_contact.state:
        if message.contact:
            target_user_id = message.contact.user_id
            if not target_user_id:
                await message.answer(
                    "⚠️ У этого контакта скрыт ID в настройках приватности Telegram.\n"
                    "Попробуйте использовать кнопку **'👤 Отметить по юзернейму'**."
                )
                return

            phone = message.contact.phone_number or ""
            name = f"{message.contact.first_name or ''} {message.contact.last_name or ''}".strip()

            upsert_user_profile(user_id=target_user_id, phone_number=phone, full_name=name)
            await state.update_data(target_user_id=target_user_id, target_name=name)
            await state.set_state(AdminCashState.waiting_for_amount)
            await message.answer(
                f"👤 Клиент: **{escape_md(name)}**\n\n💰 Введите сумму оплаты (в тенге), например: `10000`",
                reply_markup=types.ReplyKeyboardRemove()
            )
        else:
            await message.answer("⚠️ Отправьте карточку контакта человека!")
        return

    if current_state == AdminCashState.waiting_for_amount.state:
        amount = parse_amount(message.text)
        if amount is None:
            await message.answer("⚠️ Введите корректную сумму числом, например: `10000`")
            return

        data = await state.get_data()
        target_user_id = data["target_user_id"]
        target_name = data["target_name"]
        await state.clear()

        rid = _next_receipt_id()
        pending_cash_confirmations[rid] = {
            "target_user_id": target_user_id,
            "target_name": target_name,
            "amount": amount,
        }

        builder = InlineKeyboardBuilder()
        builder.button(text=f"✅ Подтвердить {format_amount(amount)} ₸", callback_data=f"cashconfirm_{rid}")
        builder.button(text="❌ Отмена", callback_data=f"cashcancel_{rid}")
        builder.adjust(1)

        await message.answer(
            f"👤 **Клиент:** {escape_md(target_name)}\n"
            f"💰 **Сумма:** {format_amount(amount)} ₸\n\n"
            f"Подтвердите оплату наличными.",
            reply_markup=builder.as_markup()
        )
        return

    if current_state == AdminUsernameState.waiting_for_username.state:
        username_input = message.text.strip()
        user_id = find_user_id_by_username(username_input)

        if not user_id:
            await state.clear()
            await message.answer(
                f"⚠️ Пользователь **{username_input}** еще не взаимодействовал с ботом в группе.",
                reply_markup=get_admin_keyboard()
            )
            return

        await state.update_data(target_user_id=user_id, target_name=username_input)
        await state.set_state(AdminUsernameState.waiting_for_amount)
        await message.answer(f"💰 Введите сумму оплаты для **{username_input}** (в тенге):")
        return

    if current_state == AdminUsernameState.waiting_for_amount.state:
        amount = parse_amount(message.text)
        if amount is None:
            await message.answer("⚠️ Введите корректную сумму числом, например: `10000`")
            return

        data = await state.get_data()
        target_user_id = data["target_user_id"]
        target_name = data["target_name"]

        record_payment(user_id=target_user_id, amount=amount, pay_type="Админ (ручной ввод)")

        # Синхронизация с Google Таблицей
        profile = get_user_profile(target_user_id) or {}
        await sync_to_google_sheet(
            full_name=profile.get("full_name") or target_name,
            phone=profile.get("phone_number") or "",
            notes=f"Оплата через админа: {amount} KZT"
        )

        await state.clear()
        await message.answer(
            f"✅ **{target_name} успешно отмечен!**\n"
            f"• **Сумма:** {format_amount(amount)} ₸\n"
            f"Статус оплаты обновлён за текущий месяц.",
            reply_markup=get_admin_keyboard()
        )
        return

    if current_state == AdminApproveState.waiting_for_amount.state:
        amount = parse_amount(message.text)
        if amount is None:
            await message.answer("⚠️ Введите корректную сумму числом, например: `10000`")
            return

        data = await state.get_data()
        target_user_id = data["target_user_id"]
        target_chat_id = data["target_chat_id"]
        caption = data["caption"]
        message_id = data["message_id"]

        await finalize_approval(
            admin_message=message,
            target_user_id=target_user_id,
            target_chat_id=target_chat_id,
            amount=amount,
            caption=caption,
            admin_message_id=message_id,
        )
        await state.clear()
        return

    if message.text == "📊 Отчёт за месяц":
        month = datetime.now().strftime("%Y-%m")
        rows = get_month_report(month)

        if not rows:
            await message.answer(f"📊 За **{month}** оплат ещё нет.")
            return

        lines = [f"📊 **Отчёт по оплатам за {month}:**\n"]
        total = 0.0
        current_user_id = None
        person_total = 0.0
        person_count = 0

        for user_id, full_name, username, amount, pay_type, paid_at in rows:
            uname = f"@{escape_md(username)}" if username else "—"
            safe_name = escape_md(full_name) if full_name else "—"

            if current_user_id is not None and user_id != current_user_id and person_count > 1:
                lines.append(f"   ↳ *Итого по клиенту: {format_amount(person_total)} ₸ ({person_count} чек(а/ов))*")

            if current_user_id != user_id:
                person_total, person_count = 0.0, 0

            lines.append(f"• {safe_name} ({uname}) — {format_amount(amount)} ₸ [{pay_type}]")
            person_total += amount
            person_count += 1
            current_user_id = user_id
            total += amount

        if person_count > 1:
            lines.append(f"   ↳ *Итого по клиенту: {format_amount(person_total)} ₸ ({person_count} чек(а/ов))*")

        unique_payers = len({r[0] for r in rows})
        lines.append(f"\n💰 **Итого:** {format_amount(total)} ₸ от {unique_payers} чел. ({len(rows)} чек(ов))")
        await message.answer("\n".join(lines))
        return

    if message.text == "📋 Список должников":
        month = datetime.now().strftime("%Y-%m")
        rows = get_unpaid_list(month)

        if not rows:
            await message.answer(f"📋 За **{month}** должников нет — все известные боту оплатили.")
            return

        lines = [f"📋 **Не оплатили за {month}:**\n"]
        for user_id, full_name, username in rows:
            uname = f"@{escape_md(username)}" if username else "нет юзернейма"
            safe_name = escape_md(full_name) if full_name else "—"
            lines.append(f"• {safe_name} ({uname}) — ID `{user_id}`")

        lines.append("\nℹ️ Список включает только тех, кто зафиксирован в базе бота.")
        await message.answer("\n".join(lines))
        return


# ================= РАБОТА В ГРУППАХ =================

def _register_group_member(message: types.Message):
    if message.from_user.is_bot:
        return
    try:
        upsert_user_profile(
            user_id=message.from_user.id,
            username=message.from_user.username or "",
            full_name=message.from_user.full_name,
            chat_id=message.chat.id,
        )
    except Exception as e:
        print(f"[Register Group Member Error]: {e}")


@dp.message(F.chat.type.in_(["group", "supergroup"]), Command("start"))
async def cmd_start_group(message: types.Message):
    _register_group_member(message)
    await message.answer(
        "👋 **Здравствуйте! Я ваш ассистент контроля оплаты DemaLike.**\n\n"
        "📌 Напоминаю, что оплата производится ежемесячно **с 1 по 7 число**.\n"
        "Нажмите кнопку ниже, чтобы оплатить абонемент или отправить чек.",
        reply_markup=get_group_keyboard()
    )


@dp.message(F.chat.type.in_(["group", "supergroup"]), F.text == "💳 Оплатить через Kaspi")
async def pay_kaspi_start_group(message: types.Message):
    _register_group_member(message)
    builder = InlineKeyboardBuilder()
    builder.button(text="📱 Открыть Kaspi Pay", url=KASPI_PAY_LINK)

    await message.answer(
        f"💳 **Оплата подписки DemaLike**\n"
        f"🗓 **Срок оплаты:** с 1 по 7 число каждого месяца.\n\n"
        f"1️⃣ Нажмите кнопку ниже для перехода в **Kaspi Pay**.\n"
        f"2️⃣ Отправьте сюда в группу **скриншот или PDF-чек** оплаты.",
        reply_markup=builder.as_markup()
    )


pending_receipts = {}
pending_cash_confirmations = {}
_receipt_counter = 0


def _next_receipt_id() -> str:
    global _receipt_counter
    _receipt_counter += 1
    return str(_receipt_counter)


@dp.message(F.chat.type.in_(["group", "supergroup"]), F.photo | F.document)
async def process_check_upload_group(message: types.Message, state: FSMContext):
    if message.from_user.is_bot:
        return

    _register_group_member(message)
    await state.clear()

    is_pdf = bool(message.document and (message.document.mime_type == "application/pdf"))
    is_image_doc = bool(message.document and (message.document.mime_type or "").startswith("image/"))
    ocr_eligible = bool(message.photo) or is_pdf or is_image_doc

    rid = _next_receipt_id()
    pending_receipts[rid] = {
        "chat_id": message.chat.id,
        "chat_title": message.chat.title,
        "message_id": message.message_id,
        "user_id": message.from_user.id,
        "username": message.from_user.username or "",
        "full_name": message.from_user.full_name,
        "photo_file_id": message.photo[-1].file_id if message.photo else None,
        "document_file_id": message.document.file_id if message.document else None,
        "detected_amount": None,
    }

    if ocr_eligible:
        file_id = pending_receipts[rid]["photo_file_id"] or pending_receipts[rid]["document_file_id"]
        detected_amount = await extract_amount_from_telegram_file(file_id, is_pdf=is_pdf)
        pending_receipts[rid]["detected_amount"] = detected_amount

    builder = InlineKeyboardBuilder()
    builder.button(text="✅ Да, это чек об оплате", callback_data=f"confirmreceipt_{rid}")
    builder.button(text="❌ Нет, это не чек", callback_data=f"ignorereceipt_{rid}")
    builder.adjust(1)

    detected = pending_receipts[rid]["detected_amount"]
    prompt = "📎 Это чек об оплате подписки DemaLike?"
    if detected:
        prompt += f"\n\n🔎 Распознанная сумма: **{format_amount(detected)} ₸**"

    await message.reply(prompt, reply_markup=builder.as_markup())


@dp.callback_query(F.data.startswith("confirmreceipt_"))
async def confirm_receipt(callback: types.CallbackQuery):
    rid = callback.data.split("_", 1)[1]
    data = pending_receipts.get(rid)

    if not data:
        await callback.answer("⚠️ Это подтверждение уже устарело или было обработано.", show_alert=True)
        return

    if callback.from_user.id != data["user_id"] and callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Подтвердить может только отправитель файла.", show_alert=True)
        return

    await callback.answer()

    detected = data["detected_amount"]
    status_text = "⏳ **Чек получен!** Передан администратору на проверку."
    if detected:
        status_text += f"\n🔎 Обнаруженная сумма: **{format_amount(detected)} ₸**"
    await callback.message.edit_text(status_text)

    current_chat_id = data["chat_id"]

    try:
        upsert_user_profile(
            user_id=data["user_id"],
            username=data["username"],
            full_name=data["full_name"],
            chat_id=current_chat_id
        )
    except Exception as e:
        print(f"[Save Profile Error]: {e}")

    for admin_id in ADMIN_IDS:
        builder = InlineKeyboardBuilder()
        if detected:
            builder.button(text=f"✅ Подтвердить {format_amount(detected)} ₸", callback_data=f"approveauto_{rid}")
            builder.button(text="✏️ Другая сумма", callback_data=f"approve_{rid}")
        else:
            builder.button(text="✅ Подтвердить", callback_data=f"approve_{rid}")
        builder.button(text="❌ Отклонить", callback_data=f"reject_{rid}")
        builder.adjust(1)

        username_str = f"@{escape_md(data['username'])}" if data["username"] else "нет юзернейма"
        safe_name = escape_md(data["full_name"])
        safe_chat_title = escape_md(data["chat_title"]) if data["chat_title"] else str(current_chat_id)
        amount_line = (
            f"• **Сумма (авто-распознавание):** {format_amount(detected)} ₸\n"
            if detected else
            "• **Сумма:** не распознана — введите вручную\n"
        )
        caption = (
            f"📥 **Новый чек на проверку (DemaLike)!**\n\n"
            f"• **Имя:** {safe_name}\n"
            f"• **Профиль:** {username_str}\n"
            f"• **ID:** `{data['user_id']}`\n"
            f"• **Группа:** `{safe_chat_title}`\n"
            f"{amount_line}"
        )

        try:
            if data["photo_file_id"]:
                await bot.send_photo(admin_id, data["photo_file_id"], caption=caption, reply_markup=builder.as_markup())
            elif data["document_file_id"]:
                await bot.send_document(admin_id, data["document_file_id"], caption=caption,
                                        reply_markup=builder.as_markup())
        except Exception as e:
            print(f"[Admin Notice Error]: {e}")


@dp.callback_query(F.data.startswith("ignorereceipt_"))
async def ignore_receipt(callback: types.CallbackQuery):
    rid = callback.data.split("_", 1)[1]
    data = pending_receipts.get(rid)

    if data and callback.from_user.id != data["user_id"] and callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Отменить может только отправитель файла.", show_alert=True)
        return

    pending_receipts.pop(rid, None)
    await callback.answer()
    await callback.message.edit_text("Хорошо, файл не будет передан администратору.")


# ================= ПОДТВЕРЖДЕНИЕ ОПЛАТЫ НАЛИЧНЫМИ =================

@dp.callback_query(F.data.startswith("cashconfirm_"))
async def cash_confirm(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        return

    rid = callback.data.split("_", 1)[1]
    data = pending_cash_confirmations.pop(rid, None)
    if not data:
        await callback.answer("⚠️ Это подтверждение уже устарело.", show_alert=True)
        return

    record_payment(user_id=data["target_user_id"], amount=data["amount"], pay_type="Наличные")

    # Синхронизация с Google Таблицей
    profile = get_user_profile(data["target_user_id"]) or {}
    await sync_to_google_sheet(
        full_name=profile.get("full_name") or data["target_name"],
        phone=profile.get("phone_number") or "",
        notes=f"Наличные: {data['amount']} KZT"
    )

    await callback.answer()
    await callback.message.edit_text(
        f"✅ **Оплата наличными зафиксирована!**\n\n"
        f"• **Клиент:** {escape_md(data['target_name'])}\n"
        f"• **Сумма:** {format_amount(data['amount'])} ₸\n"
        f"• **Статус:** оплачено за текущий месяц."
    )
    await callback.message.answer("Готово ✅", reply_markup=get_admin_keyboard())


@dp.callback_query(F.data.startswith("cashcancel_"))
async def cash_cancel(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        return

    rid = callback.data.split("_", 1)[1]
    pending_cash_confirmations.pop(rid, None)
    await callback.answer()
    await callback.message.edit_text("❌ Отменено. Сумма не записана.")
    await callback.message.answer("Главное меню:", reply_markup=get_admin_keyboard())


# ================= ПРОВЕРКА ЧЕКА АДМИНОМ =================

async def finalize_approval(admin_message, target_user_id, target_chat_id, amount,
                            caption, admin_message_id):
    record_payment(user_id=target_user_id, amount=amount, pay_type="Kaspi")

    # Синхронизация с Google Таблицей
    profile = get_user_profile(target_user_id) or {}
    await sync_to_google_sheet(
        full_name=profile.get("full_name") or "Ученик",
        phone=profile.get("phone_number") or "",
        notes=f"Оплата Kaspi: {amount} KZT"
    )

    try:
        await bot.edit_message_caption(
            chat_id=admin_message.from_user.id,
            message_id=admin_message_id,
            caption=caption + f"\n\n✅ **ОПЛАТА ПОДТВЕРЖДЕНА** — {format_amount(amount)} ₸"
        )
    except Exception as e:
        print(f"[Edit Caption Error]: {e}")

    try:
        await bot.send_message(target_chat_id, "✅ **Подтверждено!** Оплата за текущий месяц принята.")
    except Exception as e:
        print(f"[Group Send Error]: {e}")

    await admin_message.answer("Готово ✅", reply_markup=get_admin_keyboard())


@dp.callback_query(F.data.startswith("approveauto_"))
async def approve_payment_auto(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        return

    rid = callback.data.split("_", 1)[1]
    data = pending_receipts.pop(rid, None)
    if not data or not data["detected_amount"]:
        await callback.answer("⚠️ Данные по этому чеку устарели — используйте «✏️ Другая сумма».", show_alert=True)
        return

    await callback.answer()
    caption = callback.message.caption or ""
    await finalize_approval(
        admin_message=callback.message,
        target_user_id=data["user_id"],
        target_chat_id=data["chat_id"],
        amount=data["detected_amount"],
        caption=caption,
        admin_message_id=callback.message.message_id,
    )


@dp.callback_query(F.data.startswith("approve_"))
async def approve_payment(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        return

    rid = callback.data.split("_", 1)[1]
    data = pending_receipts.pop(rid, None)
    if not data:
        await callback.answer("⚠️ Это подтверждение уже устарело или было обработано.", show_alert=True)
        return

    caption = callback.message.caption or ""

    await state.set_state(AdminApproveState.waiting_for_amount)
    await state.update_data(
        target_user_id=data["user_id"],
        target_chat_id=data["chat_id"],
        caption=caption,
        message_id=callback.message.message_id
    )

    await callback.answer()
    await callback.message.reply("💰 Введите сумму оплаты по этому чеку (в тенге), например: `10000`")


@dp.callback_query(F.data.startswith("reject_"))
async def reject_payment(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        return

    rid = callback.data.split("_", 1)[1]
    data = pending_receipts.pop(rid, None)
    target_chat_id = data["chat_id"] if data else None

    await callback.message.edit_caption(
        caption=(callback.message.caption or "") + "\n\n❌ **ОПЛАТА ОТКЛОНЕНА**"
    )
    await callback.answer()

    if target_chat_id is not None:
        try:
            await bot.send_message(
                target_chat_id,
                "❌ **Чек отклонен.** Пожалуйста, свяжитесь с администратором и пришлите корректный чек."
            )
        except Exception:
            pass


# ================= ЕЖЕМЕСЯЧНЫЙ ОТЧЁТ О ДОЛЖНИКАХ =================

async def send_monthly_unpaid_report():
    month = datetime.now().strftime("%Y-%m")
    rows = get_unpaid_list(month)

    if not rows:
        text = f"📋 **Отчёт за {month}:** все известные боту пользователи оплатили. 🎉"
    else:
        lines = [f"📋 **Автоматический отчёт о должниках за {month}:**\n"]
        for user_id, full_name, username in rows:
            uname = f"@{escape_md(username)}" if username else "нет юзернейма"
            safe_name = escape_md(full_name) if full_name else "—"
            lines.append(f"• {safe_name} ({uname}) — ID `{user_id}`")
        text = "\n".join(lines)

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text)
        except Exception as e:
            print(f"[Monthly Report Error]: {e}")


# ================= ЗАПУСК И ПРИВЕТСТВИЕ В ГРУППАХ =================

async def send_startup_greeting():
    greeting_text = (
        "🤖 **Здравствуйте! Я ваш ассистент DemaLike.**\n\n"
        "Я запущен и готов к работе. Напоминаю:\n"
        "• Оплата принимается ежемесячно **с 1 по 7 число**.\n"
        "• Воспользуйтесь кнопкой **«💳 Оплатить через Kaspi»** ниже для отправки чека."
    )
    for chat_id in TARGET_CHAT_IDS:
        try:
            await bot.send_message(chat_id, greeting_text, reply_markup=get_group_keyboard())
        except Exception as e:
            print(f"[Startup Greeting Error] Chat {chat_id}: {e}")


async def main():
    init_db()

    scheduler.add_job(send_monthly_unpaid_report, "cron", day=8, hour=0, minute=1)
    scheduler.start()

    print("🚀 DemaLike Bot запущен и готов к работе!")
    print(f"OCR-автораспознавание суммы: {'включено' if OCR_AVAILABLE else 'выключено'}")

    await send_startup_greeting()

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())